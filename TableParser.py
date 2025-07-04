from openpyxl import load_workbook
from openpyxl.utils.cell import *
import json

class TableParser:
    def __init__(self, config):
        self.config = config
    
    def find_reasons(self, worksheet):
        reasons = {}
        coord = coordinate_from_string(self.config.reasons_location)
        new_coord = get_column_letter(column_index_from_string(coord[0]) + 1) + str(coord[1] + self.config.reasons_count - 1)
        for (color, reason) in worksheet[self.config.reasons_location:new_coord]:
            reasons[color.fill.fgColor.rgb] = reason.value
        return reasons
    
    def find_calendar(self, worksheet, date):
        for cell in worksheet['A']:
            if cell.data_type == 'd' and cell.value.strftime('%Y-%m') == date.strftime('%Y-%m'):
                return cell.row

    def find_colors(self, worksheet, row, date):
        colors = {}
        for i in range(self.config.dba_count):
            colors[worksheet[f'A{row + 3 + i}'].value] = worksheet[f'{get_column_letter(date.day + 1)}{row + 3 + i}'].fill.fgColor.rgb
        return colors

    def parse_reasons(self, date):
        wb = load_workbook(self.config.table_path)
        ws = wb.active

        reasons = self.find_reasons(ws)
        colors = self.find_colors(ws, self.find_calendar(ws, date), date)
        
        res = {}
        for i in colors:
            if colors[i] in reasons:
                res[i] = reasons[colors[i]]
        return res
    
    def make_phrase(self, date):
        try:
            with open('phrases.json', encoding='utf-8') as f:
                phrases = json.loads(f.read())
            phrase = phrases['hello'] + '\n\n'
            phrase += '\n'.join(phrases['reasons'][j].format(i) for i,j in self.parse_reasons(date).items())
            phrase += '\n\n' + phrases['bye']
            return phrase
        except Exception as e:
            return phrases['error'].format(e)